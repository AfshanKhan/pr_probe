from typing import List, Optional
import json
from .models import PRAnalysisResult, SummaryMetrics, RepoMetrics

def export_json(results: List[PRAnalysisResult], metrics: SummaryMetrics, filename: str):
    data = {
        "summary": metrics.model_dump(mode='json'),
        "results": [r.model_dump(mode='json') for r in results]
    }
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)

def format_duration(hours: Optional[float]) -> str:
    if hours is None:
        return "N/A"
    total_minutes = int(hours * 60)
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h}h:{m}m"

def export_xlsx(results: List[PRAnalysisResult], metrics: SummaryMetrics, filename: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    
    # 1. Detailed Report Sheet
    ws1 = wb.active
    ws1.title = "Detailed Report"
    
    if results:
        headers = list(results[0].model_dump().keys())
        ws1.append(headers)
        
        # Bold headers
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            
        for r in results:
            data = r.model_dump(mode='json')
            # Format durations for Excel
            data['tat_hours'] = format_duration(r.tat_hours)
            data['ttr_hours'] = format_duration(r.ttr_hours)
            ws1.append(list(data.values()))
            
        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws1.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # 2. Summary Metrics Sheet
    ws2 = wb.create_sheet(title="Summary Metrics")
    ws2.append(["OVERALL METRICS", ""])
    ws2["A1"].font = Font(bold=True, size=12)
    
    summary_data = [
        ["Total PRs Merged", metrics.total_prs],
        ["Template Usage Count", metrics.template_usage_count],
        ["Template Usage Percent", f"{metrics.template_usage_percent:.1f}%"],
        ["Approved Before Merge Count", metrics.approved_before_merge_count],
        ["Approval Percent", f"{metrics.approval_percent:.1f}%"],
        ["Average Turnaround Time (TAT)", format_duration(metrics.avg_tat_hours)],
        ["Average Time to 1st Review (TTR)", format_duration(metrics.avg_ttr_hours)]
    ]
    
    for row in summary_data:
        ws2.append(row)
        
    # Add Per-Repository Breakdown
    if metrics.repo_metrics:
        ws2.append([])
        ws2.append(["PER-REPOSITORY METRICS", ""])
        ws2.append(["Repository", "PRs", "Template Usage", "Approved", "Avg TAT", "Avg TTR"])
        
        # Bold the section header and table headers
        row_idx = ws2.max_row - 1
        ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
            
        for repo, m_dict in metrics.repo_metrics.items():
            # Convert dict back to model to use properties
            m = RepoMetrics(**m_dict)
            ws2.append([
                repo, 
                m.total_prs,
                f"{m.template_usage_count} ({m.template_usage_percent:.1f}%)",
                f"{m.approved_before_merge_count} ({m.approval_percent:.1f}%)",
                format_duration(m.avg_tat_hours),
                format_duration(m.avg_ttr_hours)
            ])
        
    # Style summary sheet
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 15
    
    wb.save(filename)

def calculate_metrics(results: List[PRAnalysisResult]) -> SummaryMetrics:
    total = len(results)
    if total == 0:
        return SummaryMetrics(
            total_prs=0, template_usage_count=0, approved_before_merge_count=0,
            avg_tat_hours=0.0, avg_ttr_hours=0.0, repo_metrics={}
        )
        
    template_count = sum(1 for r in results if r.template_used)
    approval_count = sum(1 for r in results if r.approved_before_merge)
    
    # Overall averages
    avg_tat = sum(r.tat_hours for r in results) / total
    ttr_values = [r.ttr_hours for r in results if r.ttr_hours is not None]
    avg_ttr = sum(ttr_values) / len(ttr_values) if ttr_values else 0.0
    
    # Per-repo averages
    repo_groups = {}
    for r in results:
        if r.repo not in repo_groups:
            repo_groups[r.repo] = []
        repo_groups[r.repo].append(r)
        
    repo_metrics = {}
    for repo, prs in repo_groups.items():
        r_total = len(prs)
        r_template = sum(1 for p in prs if p.template_used)
        r_approval = sum(1 for p in prs if p.approved_before_merge)
        r_avg_tat = sum(p.tat_hours for p in prs) / r_total
        r_ttr_values = [p.ttr_hours for p in prs if p.ttr_hours is not None]
        r_avg_ttr = sum(r_ttr_values) / len(r_ttr_values) if r_ttr_values else 0.0
        
        repo_metrics[repo] = RepoMetrics(
            total_prs=r_total,
            template_usage_count=r_template,
            approved_before_merge_count=r_approval,
            avg_tat_hours=round(r_avg_tat, 2),
            avg_ttr_hours=round(r_avg_ttr, 2)
        ).model_dump(mode='json')
    
    return SummaryMetrics(
        total_prs=total,
        template_usage_count=template_count,
        approved_before_merge_count=approval_count,
        avg_tat_hours=round(avg_tat, 2),
        avg_ttr_hours=round(avg_ttr, 2),
        repo_metrics=repo_metrics
    )
